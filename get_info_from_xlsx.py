import openpyxl

def process_xlsx(folder_name):
    theFile = openpyxl.load_workbook("./Annotation_List_Luiggi.xlsx")
    allSheetNames = theFile.sheetnames

    folder = folder_name
    metadata = []

    for sheet in allSheetNames:
        #print("Current sheet name is {}" .format(sheet))
        currentSheet = theFile[sheet]
        for row in range(1, currentSheet.max_row + 1):
            first_cell = "{}{}".format("A",row)
            if currentSheet[first_cell].value == folder:
                aux = []
                for column in "ABCDEFG":  # Columns that is going to scan.
                    cell_name = "{}{}".format(column, row)
                    #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                    aux.append(currentSheet[cell_name].value)
                metadata.append(aux)
            else:
                continue
    
    return metadata

def process_xlsx_metadata(folder_name):
    theFile = openpyxl.load_workbook("./Metadata_Luiggi.xlsx")
    allSheetNames = theFile.sheetnames

    folder = folder_name
    metadata = []

    for sheet in allSheetNames:
        #print("Current sheet name is {}" .format(sheet))
        currentSheet = theFile[sheet]
        for row in range(1, currentSheet.max_row + 1):
            first_cell = "{}{}".format("A",row)
            if currentSheet[first_cell].value == folder:
                aux = []
                for column in "ABCDEFGHIJK":  # Columns that is going to scan.
                    cell_name = "{}{}".format(column, row)
                    #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                    aux.append(currentSheet[cell_name].value)
                metadata.append(aux)
            else:
                continue
    
    return metadata

if __name__ == "__main__":
    x = process_xlsx("1CM1_1_R_#217")
    print (x)
    print(len(x))